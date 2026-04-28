"""
將 Excel 中的測驗題目和結果匯出為純文字檔。
請在終端機中執行：
  pip install openpyxl
  python extract_quiz.py
產出的 quiz_data.txt 檔案會在同一個資料夾中。
"""
import openpyxl, os

xlsx_path = os.path.join(os.path.dirname(__file__), "異世界轉生_異世界轉生：【靈魂軌跡】故事流.xlsx")
wb = openpyxl.load_workbook(xlsx_path)

out_path = os.path.join(os.path.dirname(__file__), "quiz_data.txt")
with open(out_path, "w", encoding="utf-8") as f:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write(f"=== Sheet: {sheet_name} ===\n")
        f.write(f"Rows: {ws.max_row}, Cols: {ws.max_column}\n\n")
        for i, row in enumerate(ws.iter_rows(), 1):
            values = [str(cell.value) if cell.value is not None else "" for cell in row]
            f.write(f"Row {i}: {' | '.join(values)}\n")
        f.write("\n\n")

print(f"Done! Output saved to: {out_path}")
